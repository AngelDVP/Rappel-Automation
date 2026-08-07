"""
analisis_rapel.py
-----------------
Procesa facturas de acuerdos comerciales (Rapel) mediante visión de IA y cruce 
de datos con la BBDD de provisiones anuales para sugerir el mes de carga contable.
"""

import sys
import os
import io
import re
import base64
import unicodedata
import openpyxl
from datetime import datetime
from openai import OpenAI
import pypdfium2 as pdfium
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuración
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY") or ""
OPENAI_MODEL   = "gpt-4o"  

PROTOCOLO_FILE = "Protocolo Rapel.xlsx"
BBDD_FILE      = "BBDD RAPEL.xlsx"

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}
MESES_NOMBRE = {v: k.capitalize() for k, v in MESES_ES.items()}


def quitar_acentos(s):
    """Normaliza texto eliminando acentos."""
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')


def normalizar_rut(rut):
    """Extrae números del RUT y formatea a 8 dígitos."""
    if not rut:
        return None
    s = str(rut).strip()
    if "-" in s:
        s = s.split("-")[0]
    s = s.replace(".", "")
    return s.zfill(8)


def detectar_mes_glosa(glosa):
    """Extrae mes y año mencionados en la glosa de la factura."""
    if not glosa:
        return None, None
    lower = glosa.lower()
    for nombre, num in MESES_ES.items():
        if nombre in lower:
            match = re.search(r"\b(20\d{2})\b", glosa)
            year = int(match.group(1)) if match else datetime.now().year
            return num, year
    return None, None


def antiguedad_meses(fecha):
    """Calcula diferencia en meses respecto a la fecha actual."""
    if not fecha:
        return 0
    hoy = datetime.now()
    return max(0, (hoy.year - fecha.year) * 12 + (hoy.month - fecha.month))


def find_pdf(folder, n_doc):
    """Busca un archivo PDF que contenga el número de documento."""
    num_str = str(n_doc)
    for fname in os.listdir(folder):
        if fname.lower().endswith(".pdf") and num_str in fname:
            return os.path.join(folder, fname)
    return None


def pdf_a_imagen(pdf_path, scale=1.5):
    """Renderiza la primera página y recorta el bloque de la descripción."""
    doc = pdfium.PdfDocument(pdf_path)
    page = doc[0]
    bitmap = page.render(scale=scale)
    img = bitmap.to_pil()
    doc.close()
    w, h = img.size
    img_crop = img.crop((0, int(h * 0.28), int(w * 0.80), int(h * 0.58)))
    buf = io.BytesIO()
    img_crop.save(buf, format="PNG")
    return buf.getvalue()


def extraer_glosa_vision(client, img_bytes, n_doc):
    """Consulta OpenAI Vision para OCR del bloque de descripción."""
    img_b64 = base64.b64encode(img_bytes).decode("utf-8")
    prompt = (
        f"Esta imagen es una seccion de una factura electronica chilena N {n_doc}. "
        "Extrae UNICAMENTE el texto de la columna 'Descripcion' de la tabla de productos/servicios. "
        "Responde SOLO con ese texto, sin comillas ni explicaciones. "
        "Si hay varias lineas, separalas con ' | '. "
        "Si no puedes leer nada, responde exactamente: NO_ENCONTRADO"
    )
    resp = client.chat.completions.create(
        model=OPENAI_MODEL,
        messages=[{"role": "user", "content": [
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {
                "url": f"data:image/png;base64,{img_b64}", "detail": "low"
            }}
        ]}],
        max_tokens=200, temperature=0
    )
    return resp.choices[0].message.content.strip()


def cargar_protocolo(script_dir):
    """Carga los parámetros del archivo maestro de reglas de negocio."""
    path = os.path.join(script_dir, PROTOCOLO_FILE)
    wb = openpyxl.load_workbook(path)

    categorizacion = []
    for row in wb["Categorizacion"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                cuenta_val = str(int(row[1])) if row[1] and str(row[1]).strip() not in ("", "-") else ""
            except (ValueError, TypeError):
                cuenta_val = ""
            categorizacion.append({
                "categoria": str(row[0]).strip(),
                "cuenta":    cuenta_val,
                "keyword":   str(row[2]).strip().lower() if row[2] else ""
            })

    prioridades = []
    for row in wb["Prioridades"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            prioridades.append({
                "cuenta":      str(int(row[0])),
                "categoria":   str(row[1]).strip().upper() if row[1] else "",
                "antiguedad":  str(row[2]).strip().upper() if row[2] else "",
                "sin_gasto":   str(row[3]).strip() if row[3] else "",
                "con_pos":     str(row[4]).strip() if row[4] else "",
                "con_neg":     str(row[5]).strip() if row[5] else ""
            })

    num_art = []
    for row in wb["Numero articulo"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            formato = str(row[2]).strip()
            sufijo = formato.split("_")[-1].replace("'", "").replace('"', "")
            num_art.append({
                "cuenta":    str(int(row[0])),
                "categoria": str(row[1]).strip(),
                "sufijo":    sufijo
            })

    return categorizacion, prioridades, num_art


def cargar_bbdd(script_dir):
    """Carga la base de datos de provisiones en memoria."""
    path = os.path.join(script_dir, BBDD_FILE)
    wb = openpyxl.load_workbook(path)
    ws = wb["Hoja1"]

    bbdd = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rapel = float(row[7]) if row[7] else 0.0
        suma  = float(row[8]) if row[8] else 0.0
        dif   = float(row[9]) if row[9] else rapel - suma
        bbdd.append({
            "rut":    str(row[0]).strip().zfill(8),
            "nombre": str(row[1]).strip() if row[1] else "",
            "canal":  str(row[2]).strip() if row[2] else "",
            "year":   int(str(row[4]).strip()) if row[4] else 0,
            "month":  int(row[5]) if row[5] else 0,
            "rapel":  rapel,
            "suma":   suma,
            "dif":    dif,
        })
    return bbdd


def cargar_memo(week_folder):
    """Carga la planilla de control de facturas de la semana."""
    memo_path = None
    for f in os.listdir(week_folder):
        if f.lower().endswith(".xlsx") and not f.lower().startswith("analisis"):
            memo_path = os.path.join(week_folder, f)
            break
    if not memo_path:
        raise FileNotFoundError(f"No se encontró memo Excel en: {week_folder}")

    wb = openpyxl.load_workbook(memo_path)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any("documento" in str(v).lower() for v in row if v):
            header_row = i
            break
    if not header_row:
        raise ValueError("No se encontró la fila de encabezados en el memo.")

    facturas = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[4]:  
            continue
        facturas.append({
            "rut":    normalizar_rut(row[2]),
            "nombre": str(row[3]).strip() if row[3] else "",
            "n_doc":  int(row[4]),
            "fecha":  row[5] if isinstance(row[5], datetime) else None,
            "monto":  float(row[6]) if row[6] else 0.0,
        })
    return facturas


def categorizar(glosa, categorizacion):
    """Categoriza una glosa comparando coincidencia de palabras clave."""
    if not glosa or glosa == "NO_ENCONTRADO":
        return "Sin glosa", ""
    glosa_norm = quitar_acentos(glosa).lower().replace("rappel", "rapel")

    for entry in categorizacion:
        if not entry["keyword"]:
            continue
        keywords = [quitar_acentos(k).lower().strip() for k in entry["keyword"].split(",") if k.strip()]
        if all(kw in glosa_norm for kw in keywords):
            return entry["categoria"], entry["cuenta"]

    return "Sin categoria", ""


def get_numero_articulo(cuenta, categoria, mes, year, num_art_table):
    """Genera código del artículo formateado: YYYY_MM_SUFIJO."""
    if not mes or not year:
        return ""
    cat_clean = categoria.replace(" Independiente", "").replace(" independiente", "").strip()
    for entry in num_art_table:
        if entry["cuenta"] == cuenta or entry["categoria"].lower() == cat_clean.lower():
            return f"{year}_{str(mes).zfill(2)}_{entry['sufijo']}"
    return ""


def aplicar_regla(accion_str):
    """Valida reglas específicas del protocolo y retorna (aplica, alerta)."""
    norm = quitar_acentos(accion_str).lower().strip()
    asignar = norm.startswith("si")
    match = re.search(r'["\u201c\u201d]([^"\u201c\u201d]+)["\u201c\u201d]', accion_str)
    alerta = match.group(1) if match else None
    if not asignar and not alerta:
        alerta = "Excede"
    return asignar, alerta


def analizar_rapel(factura, bbdd, prioridades, num_art_table, cuenta, categoria):
    """Calcula y actualiza la provisión en base a prioridades y antigüedad."""
    rut      = factura["rut"]
    monto    = factura["monto"]
    fecha    = factura["fecha"]
    glosa    = factura.get("glosa", "")
    antig    = antiguedad_meses(fecha)
    year_hoy = datetime.now().year

    filas = [r for r in bbdd if r["rut"] == rut and r["year"] == year_hoy]
    if not filas:
        return "", None, None, None, None, "Cliente no encontrado en BBDD", ""

    canal = filas[0]["canal"]
    cat_up = categoria.upper()

    antig_key = "0-2 MESES" if antig <= 2 else "3+ MESES"
    cat_key   = "RAPEL INDEPENDIENTE" if "INDEPENDIENTE" in cat_up else "RAPEL"
    rule = None
    for r in prioridades:
        if r["categoria"] == cat_key:
            if r["antiguedad"] == "CUALQUIERA" or r["antiguedad"] == antig_key:
                rule = r
                break

    if "INDEPENDIENTE" in cat_up:
        mes_g, year_g = detectar_mes_glosa(glosa)
        fila = next((r for r in filas if r["month"] == mes_g), filas[0]) if mes_g else filas[0]
        prov = fila["dif"] if fila["dif"] > 0 else None
        prov_rest = (fila["dif"] - monto) if fila else None
        num_art = get_numero_articulo(cuenta, categoria, fila["month"], fila["year"], num_art_table) if fila else ""
        return canal, fila["month"], fila["year"], prov, prov_rest, 'Alerta "Independiente"', num_art

    mes_g, year_g = detectar_mes_glosa(glosa)

    if mes_g:
        fila = next((r for r in filas if r["month"] == mes_g and r["year"] == (year_g or year_hoy)), None)
        if not fila:
            return canal, mes_g, year_g or year_hoy, None, None, "Mes de glosa no encontrado en BBDD", ""

        prov_disp = fila["dif"] if fila["dif"] > 0 else None
        prov_rest = fila["dif"] - monto
        num_art   = get_numero_articulo(cuenta, categoria, fila["month"], fila["year"], num_art_table)

        if fila["suma"] == 0 or fila["dif"] > 0:
            fila["suma"] += monto   
            fila["dif"]  -= monto
            if antig >= 3:
                estado_final = 'Alerta "Rapel Antiguo"'
            elif prov_rest < 0:
                estado_final = 'Alerta "Supera Provisi\u00f3n"'
            else:
                estado_final = "OK"
            return canal, fila["month"], fila["year"], prov_disp, prov_rest, estado_final, num_art
        else:
            estado_glosa = 'Alerta "Excede"' if prov_rest < 0 else "OK"
            return canal, fila["month"], fila["year"], prov_disp, prov_rest, estado_glosa, num_art

    else:
        hoy = datetime.now()
        mes_actual_key = (hoy.year, hoy.month)

        asignables_directos = []  
        asignables_alerta   = []  
        no_asignables       = []  

        for r in filas:
            if (r["year"], r["month"]) == mes_actual_key:
                continue  
            antig_r = (hoy.year - r["year"]) * 12 + (hoy.month - r["month"])
            if r["suma"] == 0 or r["dif"] > 0:
                if antig_r <= 2:
                    asignables_directos.append(r)
                else:
                    asignables_alerta.append(r)
            elif r["dif"] <= 0 and r["suma"] > 0:
                no_asignables.append(r)

        asignables_directos.sort(key=lambda r: (r["year"], r["month"]))
        asignables_alerta.sort(key=lambda r: (r["year"], r["month"]), reverse=True)
        no_asignables.sort(key=lambda r: (r["year"], r["month"]), reverse=True)

        if not asignables_directos and not asignables_alerta and not no_asignables:
            return canal, None, None, None, None, "Sin provision en BBDD", ""

        if asignables_directos:
            fila_elegida = asignables_directos[0]
            estado_grupo = None  
        elif asignables_alerta:
            fila_elegida = asignables_alerta[0]
            estado_grupo = "antiguo"
        else:
            fila_elegida = no_asignables[0]
            prov_disp = None
            prov_rest = fila_elegida["dif"] - monto
            num_art = get_numero_articulo(cuenta, categoria, fila_elegida["month"], fila_elegida["year"], num_art_table)
            estado_excede = 'Alerta "Excede"' if prov_rest < 0 else "OK"
            return canal, fila_elegida["month"], fila_elegida["year"], prov_disp, prov_rest, estado_excede, num_art

        prov_disp = fila_elegida["dif"] if fila_elegida["dif"] > 0 else None
        prov_rest = fila_elegida["dif"] - monto

        fila_elegida["suma"] += monto   
        fila_elegida["dif"]  -= monto

        num_art = get_numero_articulo(cuenta, categoria, fila_elegida["month"], fila_elegida["year"], num_art_table)

        if estado_grupo == "antiguo":
            estado_final = 'Alerta "Rapel Antiguo"'
        elif prov_rest < 0:
            estado_final = 'Alerta "Supera Provisión"'
        else:
            estado_final = "OK"
        return canal, fila_elegida["month"], fila_elegida["year"], prov_disp, prov_rest, estado_final, num_art


def _aplicar_a_fila(fila, monto, rule, canal, cuenta, categoria, num_art_table):
    tiene_gasto = fila["suma"] > 0
    prov_pos    = fila["dif"] > 0

    if rule is None:
        accion_str = "Si"
    elif not tiene_gasto:
        accion_str = rule["sin_gasto"]
    elif prov_pos:
        accion_str = rule["con_pos"]
    else:
        accion_str = rule["con_neg"]

    asignar, alerta = aplicar_regla(accion_str)
    estado = f'Alerta "{alerta}"' if alerta else ("OK" if asignar else "No asignable")

    prov_disp  = fila["dif"] if fila["dif"] > 0 else None
    prov_rest  = fila["dif"] - monto

    if asignar:
        fila["suma"] += monto   
        fila["dif"]  -= monto

    num_art = get_numero_articulo(cuenta, categoria, fila["month"], fila["year"], num_art_table)
    return canal, fila["month"], fila["year"], prov_disp, prov_rest, estado, num_art


# Estilos de Reporte Excel
FILL_HEADER    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_OK        = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ALERTA    = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
FILL_ERROR     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GLOSA_MES = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  
FONT_HDR       = Font(color="FFFFFF", bold=True)


def guardar_output(resultados, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis"

    headers = [
        "N° Factura", "RUT Proveedor", "Nombre Proveedor", "Canal",
        "Fecha Emision", "Monto Neto", "Glosa Extraida", "Categoria",
        "Cuenta", "Mes Sugerido", "N° Articulo",
        "Provision Disponible", "Provision Restante", "Estado"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = Alignment(horizontal="center")

    for r in resultados:
        mes_str = ""
        if r.get("mes") and r.get("year_carga"):
            mes_str = f"{MESES_NOMBRE.get(r['mes'], r['mes'])} {r['year_carga']}"

        fecha_str = r["fecha"].strftime("%d-%m-%Y") if r.get("fecha") else ""

        fila = [
            r["n_doc"], r["rut"], r["nombre"], r.get("canal", ""),
            fecha_str, r["monto"], r.get("glosa", ""), r.get("categoria", ""),
            r.get("cuenta", ""), mes_str, r.get("num_art", ""),
            r.get("prov_disp"), r.get("prov_rest"), r.get("estado", "")
        ]
        ws.append(fila)

        estado = str(r.get("estado", "")).lower()
        if "alerta" in estado or any(x in estado for x in ["rapel antiguo", "independiente", "excede", "supera"]):
            fill = FILL_ALERTA
        elif any(x in estado for x in ["no encontrado", "sin provision", "sin categoria", "sin glosa", "pdf no encontrado"]):
            fill = FILL_ERROR
        elif estado == "ok" and r.get("tiene_mes_glosa"):
            fill = FILL_GLOSA_MES
        elif estado == "ok":
            fill = FILL_OK
        else:
            fill = FILL_ALERTA  
        for cell in ws[ws.max_row]:
            cell.fill = fill

    anchos = [13, 15, 30, 8, 14, 14, 55, 22, 10, 18, 13, 20, 20, 25]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    for row in ws.iter_rows(min_row=2):
        for cell in [row[11], row[12]]:  
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Uso: python analisis_rapel.py \"Semana 24\\1er envio\"")
        sys.exit(1)

    script_dir  = os.path.dirname(os.path.abspath(__file__))
    week_folder = os.path.join(script_dir, sys.argv[1])

    if not os.path.isdir(week_folder):
        print(f"Error: no existe la carpeta '{week_folder}'")
        sys.exit(1)

    folder_name = os.path.basename(os.path.normpath(week_folder))
    parent_name = os.path.basename(os.path.dirname(os.path.normpath(week_folder)))
    output_path = os.path.join(week_folder, f"Analisis {parent_name} - {folder_name}.xlsx")

    print("=" * 70)
    print("  ANALISIS RAPEL - PROCESAMIENTO DE CARGA")
    print(f"  Carpeta: {week_folder}")
    print("=" * 70)

    print("\nCargando referencias...", end=" ")
    categorizacion, prioridades, num_art_table = cargar_protocolo(script_dir)
    print(f"{len(categorizacion)} reglas cargadas.")

    print("Cargando bases de datos...", end=" ")
    bbdd = cargar_bbdd(script_dir)
    print(f"{len(bbdd)} registros en provisiones.")

    print("Cargando memo de control...", end=" ")
    facturas = cargar_memo(week_folder)
    print(f"{len(facturas)} documentos a procesar.\n")

    if not OPENAI_API_KEY:
        raise ValueError("API Key de OpenAI no configurada en las variables de entorno.")

    client = OpenAI(api_key=OPENAI_API_KEY)
    resultados = []
    total = len(facturas)

    for i, f in enumerate(facturas, 1):
        n_doc = f["n_doc"]
        print(f"  [{i:02d}/{total}] Fac {n_doc:<12}", end=" ", flush=True)

        pdf_path = find_pdf(week_folder, n_doc)
        if not pdf_path:
            print("-> PDF no encontrado")
            resultados.append({**f, "glosa": "", "categoria": "PDF no encontrado",
                                "cuenta": "", "canal": "", "mes": None, "year_carga": None,
                                "num_art": "", "prov_disp": None, "prov_rest": None,
                                "estado": "PDF no encontrado"})
            continue

        try:
            img = pdf_a_imagen(pdf_path)
            glosa = extraer_glosa_vision(client, img, n_doc)
        except Exception as e:
            glosa = "NO_ENCONTRADO"
            print(f"(error vision: {e}) ", end="", flush=True)

        f["glosa"] = glosa
        print(f"| {glosa[:35]:<35}", end=" ", flush=True)

        categoria, cuenta = categorizar(glosa, categorizacion)
        print(f"| {categoria:<22}", end=" ", flush=True)

        mes_glosa, _ = detectar_mes_glosa(glosa)
        tiene_mes_glosa = mes_glosa is not None

        if "rapel" in categoria.lower():
            canal, mes, year_c, prov_disp, prov_rest, estado, num_art = analizar_rapel(
                f, bbdd, prioridades, num_art_table, cuenta, categoria
            )
        else:
            canal, mes, year_c = "", None, None
            prov_disp, prov_rest = None, None
            num_art = ""
            estado = categoria  

        print(f"-> {estado}")

        resultados.append({**f, "glosa": glosa, "categoria": categoria,
                            "cuenta": cuenta, "canal": canal, "mes": mes,
                            "year_carga": year_c, "num_art": num_art,
                            "prov_disp": prov_disp, "prov_rest": prov_rest,
                            "estado": estado, "tiene_mes_glosa": tiene_mes_glosa})

    print(f"\nGuardando reporte final...")
    guardar_output(resultados, output_path)

    ok      = sum(1 for r in resultados if r["estado"].lower() == "ok")
    alertas = sum(1 for r in resultados if "alerta" in r["estado"].lower())
    otros   = total - ok - alertas

    print("=" * 70)
    print(f"  OK (asignados directamente):  {ok}")
    print(f"  Alertas (revision requerida): {alertas}")
    print(f"  Otros (no-Rapel / errores):   {otros}")
    print(f"\n  Reporte guardado en: {output_path}")
    print("=" * 70)


if __name__ == "__main__":
    main()
