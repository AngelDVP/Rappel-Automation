"""
extraer_oc.py
-------------
Procesa correos de Outlook Desktop filtrando por remitente y asunto
para extraer códigos RQ y OC, sincronizándolos con la base de datos de Ingreso RQ.
"""

import re
from datetime import datetime, timedelta
import sys
import win32com.client
import openpyxl
from pathlib import Path

# Configuración
EXCEL_PATH = Path(r"C:\Users\avillalobos\OneDrive - Trendy\Escritorio\Rapel\2026\Ingreso RQ.xlsx")
SHEET_NAME  = None          
COL_RQ      = "O"          
COL_OC      = "P"          
FILA_INICIO = 2

REMITENTE_FILTRO = "no-reply@trendy.cl"
ASUNTO_FILTRO    = "Nueva Orden de Compra"
CARPETA_INBOX = 6

def extraer_pares_desde_outlook():
    """Conecta a Outlook y extrae pares {RQ: OC} de los correos recibidos."""
    print("Conectando a Outlook...")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mapi    = outlook.GetNamespace("MAPI")
    except Exception as e:
        print(f"[ERROR] No se pudo conectar a Outlook: {e}")
        sys.exit(1)

    inbox    = mapi.GetDefaultFolder(CARPETA_INBOX)
    mensajes = inbox.Items
    mensajes.Sort("[ReceivedTime]", True)

    hace_un_mes = (datetime.now() - timedelta(days=30)).strftime("%m/%d/%Y")
    filtro = f"[ReceivedTime] >= '{hace_un_mes}'"
    mensajes = mensajes.Restrict(filtro)

    pares = {}  
    print(f"Buscando correos de '{REMITENTE_FILTRO}' con asunto '{ASUNTO_FILTRO}'...")

    count = 0
    for msg in mensajes:
        try:
            remitente = getattr(msg, "SenderEmailAddress", "") or ""
            asunto    = getattr(msg, "Subject", "") or ""
            cuerpo    = getattr(msg, "Body", "") or ""
        except Exception:
            continue

        if REMITENTE_FILTRO.lower() not in remitente.lower():
            continue
        if ASUNTO_FILTRO.lower() not in asunto.lower():
            continue

        rq_match = re.search(r"RQ\d+", cuerpo)
        oc_match = re.search(r"OC\d+", cuerpo)

        if rq_match and oc_match:
            rq = rq_match.group(0).strip()
            oc = oc_match.group(0).strip()

            if rq not in pares:  
                pares[rq] = oc
                count += 1

    print(f"  → {count} pares RQ-OC encontrados en correos.")
    return pares


def actualizar_excel(pares):
    """Escribe los pares OC en la columna P buscando por el RQ correspondiente."""
    if not EXCEL_PATH.exists():
        print(f"[ERROR] No se encontró el archivo: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Abriendo Excel: {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    encabezado_oc = ws[f"{COL_OC}1"]
    if encabezado_oc.value is None:
        encabezado_oc.value = "OC"
        print("  → Encabezado 'OC' agregado en columna P.")

    actualizados  = 0
    sin_par       = 0

    for fila in range(FILA_INICIO, ws.max_row + 1):
        celda_rq = ws[f"{COL_RQ}{fila}"]

        if celda_rq.value is None or str(celda_rq.value).strip() == "":
            continue

        rq_valor = str(celda_rq.value).strip()

        if rq_valor in pares:
            ws[f"{COL_OC}{fila}"] = pares[rq_valor]
            actualizados += 1
        else:
            print(f"  [!] Sin OC para: {rq_valor} (fila {fila})")
            sin_par += 1

    wb.save(EXCEL_PATH)
    wb.close()
    print(f"\n✔ Proceso finalizado. {actualizados} registros actualizados.")


if __name__ == "__main__":
    print("=" * 50)
    print("  Extractor RQ → OC")
    print("=" * 50)

    pares = extraer_pares_desde_outlook()

    if not pares:
        print("[AVISO] No se encontraron pares RQ-OC.")
        sys.exit(0)

    print("\nPares encontrados:")
    for rq, oc in sorted(pares.items()):
        print(f"  {rq}  →  {oc}")

    actualizar_excel(pares)
